from openpyxl import load_workbook
import pandas as pd
import os

def extract_hyperlinks_from_excel(excel_path, header_name='URL'):
    wb = load_workbook(filename=excel_path)
    sheet = wb.active
    header_row = next(sheet.iter_rows(values_only=True))
    column_letter = None
    for idx, value in enumerate(header_row):
        if value == header_name:
            column_letter = chr(65 + idx) 
            break
    if not column_letter:
        raise ValueError(f"Header '{header_name}' not found in the Excel sheet.")
    
    urls = []
    for row in range(2, sheet.max_row + 1):  
        cell = sheet[f'{column_letter}{row}']
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
    return urls
